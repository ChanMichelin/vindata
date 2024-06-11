import streamlit as st
import pandas as pd
import requests
import os
import openpyxl
import numpy as np
import json
from datetime import datetime
from io import BytesIO

@st.cache_data
def confirm_vin(file_path):
    
    # Some Excel files have more than 1 sheet, handle Excel files with more than 1 sheet
    wb = openpyxl.load_workbook(file_path)
    res = len(wb.sheetnames)
    if res > 1:
        raw_vin_data = pd.read_excel(file_path, 'Vehicle & Asset List', header=3)
    else:
        raw_vin_data = pd.read_excel(file_path, header=3)
        
    # Assign new column names to standardize raw_vin_data
    for column in raw_vin_data.columns:
        if 'vehicle asset name' in column.lower():
            raw_vin_data.rename(columns={column: 'Vehicle Asset Name'}, inplace=True)
        elif 'model year' in column.lower():
            raw_vin_data.rename(columns={column: 'Model Year'}, inplace=True)
        elif 'make' in column.lower():
            raw_vin_data.rename(columns={column: 'Make'}, inplace=True)
        elif 'model' in column.lower():
            raw_vin_data.rename(columns={column: 'Model'}, inplace=True)
        elif 'vin' in column.lower():
            raw_vin_data.rename(columns={column: 'VIN'}, inplace=True)
        elif 'fuel type' in column.lower():
            raw_vin_data.rename(columns={column: 'Fuel Type'}, inplace=True)
    
    # Create base URL for VIN query
    base_url = 'https://vpic.nhtsa.dot.gov/api/vehicles/DecodeVin/'
    
    # Create dataframe for CAN compatibility check
    vin_data = pd.DataFrame({'VRN': [], 'VIN': [], 'YEAR': [], 
                             'MAKE': [], 'MODEL': [], 'FUEL': [], 'COUNTRY': []})
    
    # Write relevant info into vin_data dataframe
    for ind in raw_vin_data.index:
        if pd.isna(raw_vin_data['VIN'][ind]) == False:
            vin_data.loc[ind] = [raw_vin_data['Vehicle Asset Name'][ind],
                             raw_vin_data['VIN'][ind], raw_vin_data['Model Year'][ind],
                             raw_vin_data['Make'][ind], raw_vin_data['Model'][ind],
                             raw_vin_data['Fuel Type'][ind],
                             'US']
    
    # Reset the vin dataframe index
    vin_data.reset_index(drop=True, inplace=True)
    
    # Replace NAN/NULL values with an empty string
    vin_data.replace(np.nan, '', inplace=True)
    
    # Change the values in vin_data dataframe to strings
    vin_data = vin_data.astype(str)
    
    # Create list to store dictionaries
    results = []
    
    # Extract VINs from vin_data dataframe into a list of values
    values = vin_data['VIN'].values.tolist()
    
    # Create variable to keep track of which index is being used
    ind = 0
    
    # Query the NHTSA VIN database using each VIN from the original sales document
    for value in values:
        value = str(value).replace(" ", "")
        url = base_url + value + '?format=json'
        response = requests.get(url, verify=False)
        
        try:
            data = response.json()
            decoded_values = {item['Variable']: item['Value'] for item in data['Results']}
            results.append({
                'VRN': vin_data['VRN'][ind],
                'VIN': value,
                'YEAR': decoded_values.get('Model Year', 'N/A'),
                'MAKE': decoded_values.get('Make', 'N/A'),
                'MODEL': decoded_values.get('Model', 'N/A'),
                'FUEL': decoded_values.get('Fuel Type - Primary', 'N/A'),
                'COUNTRY': 'US',
                'VEHICLE TYPE': decoded_values.get('Vehicle Type', 'N/A'),
                'ERROR CODE': decoded_values.get('Error Text', 'N/A')
            })
            ind += 1
        except json.JSONDecodeError as e:
            results.append({
                'VRN': vin_data['VRN'][ind],
                'VIN': value,
                'YEAR': 'Error',
                'MAKE': 'Error',
                'MODEL': 'Error',
                'FUEL': 'Error',
                'COUNTRY': 'Error',
                'VEHICLE TYPE': 'Error',
                'ERROR CODE': 'Error: No information found for input VIN'
            })
            ind += 1
        except requests.exceptions.Timeout as e:
            return "Timed out"
    
    # Create dataframe from list of dictionaries
    results = pd.DataFrame(results)
    
    # Create valid_vins dataframe
    valid_vins = results[~results.FUEL.isin(['Not Applicable', 'Error', None])]
    valid_vins = valid_vins[datetime.now().year - valid_vins['YEAR'].astype(int) < 30]
    valid_vins.drop(['VEHICLE TYPE', 'ERROR CODE'], axis=1, inplace=True)
    valid_vins.drop_duplicates(subset=['VIN'], inplace=True)
    
    check_list = []
    vins_checked = []
    valid_vin_list = valid_vins['VIN'].values.tolist()
    vin_data = pd.concat([vin_data, results['VEHICLE TYPE']], axis=1)
    
    for ind in vin_data.index:
        if vin_data['VIN'][ind].replace(" ", "") in valid_vin_list and vin_data['VIN'][ind] not in vins_checked:
            check_list.append('NO')
        elif vin_data['VEHICLE TYPE'][ind] == 'TRAILER':
            check_list.append('NO')
        elif 'trailer' in vin_data['MODEL'][ind].lower() or 'trailer' in vin_data['VRN'][ind].lower():
            check_list.append('NO')
        elif 'lift' in vin_data['MODEL'][ind].lower() or 'lift' in vin_data['VRN'][ind].lower():
            check_list.append('NO')
        elif 'example' in vin_data['VIN'][ind].lower():
            check_list.append('NO')
        elif vin_data['VIN'][ind] in vins_checked:
            check_list.append('YES: Duplicate Vin')
        else:
            check_list.append('YES')
        vins_checked.append(vin_data['VIN'][ind])
        
    for ind in vin_data.index:
        if vin_data['VEHICLE TYPE'][ind] == None or vin_data['VEHICLE TYPE'][ind] == 'Error':
                if 'trailer' in vin_data['MODEL'][ind].lower():
                    vin_data['VEHICLE TYPE'][ind] = 'TRAILER'
                elif 'lift' in vin_data['MODEL'][ind].lower():
                    vin_data['VEHICLE TYPE'][ind] = 'LIFT'
                elif vin_data['VEHICLE TYPE'][ind] == 'Error':
                    vin_data['VEHICLE TYPE'][ind] = 'UNKNOWN'

    vin_data['MANUAL CHECK NEEDED'] = check_list
    vin_data = pd.concat([vin_data, results['ERROR CODE']], axis=1)
    
    CAN_file_path = os.path.splitext(file_path)[0] + "_CAN.csv"
    pd.DataFrame(valid_vins).to_csv(CAN_file_path, index=False)
    
    processed_file_path = os.path.splitext(file_path)[0] + "_processed.xlsx"
    
    with pd.ExcelWriter(processed_file_path, engine='openpyxl') as writer:
        vin_data.to_excel(writer, index=False, sheet_name='Processed VINs')
        workbook = writer.book
        worksheet = writer.sheets['Processed VINs']
        
        for idx, column in enumerate(worksheet.columns):
            if worksheet.cell(row=1, column=idx + 1).value != 'ERROR CODE':
                max_length = 0
                for cell in column:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[chr(65 + idx)].width = adjusted_width
            if worksheet.cell(row=1, column=idx + 1).value == 'ERROR CODE':
                worksheet.column_dimensions[chr(65 + idx)].width = 12
        
    return processed_file_path, CAN_file_path

# Load the CSS file
def load_css():
    css_file_path = "style.css"
    if os.path.exists(css_file_path):
        with open(css_file_path) as f:
            css_content = f.read()
            st.markdown('<style>{}</style>'.format(css_content), unsafe_allow_html=True)
            print("CSS loaded successfully")  # Debugging print statement
    else:
        st.error("CSS file not found")
        print("CSS file not found")  # Debugging print statement

# Add the header with the logo
def add_header():
    st.markdown(
        """
        <header>
            <img src="static/michelin-logo.png" alt="Michelin">
            <h1>Michelin Connected Fleet</h1>
        </header>
        """,
        unsafe_allow_html=True
    )

# Add the footer
def add_footer():
    st.markdown(
        """
        <footer>
            <p>Powered by Michelin</p>
        </footer>
        """,
        unsafe_allow_html=True
    )

# Main function
def main():
    load_css()
    add_header()
    st.title("VIN Decoder")

    uploaded_file = st.file_uploader("Choose an Excel or CSV file", type=["xls", "xlsx", "csv"])

    if "processed_file_path" not in st.session_state:
        st.session_state["processed_file_path"] = None
        st.session_state["can_file_path"] = None

    if uploaded_file is not None:
        with st.spinner('Processing...'):
            input_file_path = uploaded_file.name
            with open(input_file_path, "wb") as f:
                f.write(uploaded_file.getbuffer())
            processed_file_path, can_file_path = confirm_vin(input_file_path)
            st.session_state["processed_file_path"] = processed_file_path
            st.session_state["can_file_path"] = can_file_path
            st.success('File successfully processed!')

    if st.session_state["processed_file_path"] and st.session_state["can_file_path"]:
        with open(st.session_state["processed_file_path"], "rb") as f:
            processed_data = f.read()
        with open(st.session_state["can_file_path"], "rb") as f:
            can_data = f.read()
        
        st.download_button(
            label="Download Processed File",
            data=BytesIO(processed_data),
            file_name=os.path.basename(st.session_state["processed_file_path"]),
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        st.download_button(
            label="Download CAN File",
            data=BytesIO(can_data),
            file_name=os.path.basename(st.session_state["can_file_path"]),
            mime='text/csv'
        )

    add_footer()

if __name__ == "__main__":
    main()
